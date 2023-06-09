# import module
import openpyxl
from openpyxl.styles import PatternFill

# load excel with its path
wrkbk = openpyxl.load_workbook("roster.xlsx")

sh = wrkbk.active

fill_cell1 = PatternFill(patternType='solid', fgColor='FFFF00')
fill_cell2 = PatternFill(patternType='solid', fgColor='ff0000')

# iterate through excel and display data
for i in range(1, sh.max_row+1):
	for j in range(1, sh.max_column+1):
		cell_obj = sh.cell(row=i, column=j)
		if cell_obj.value=='as':
		    cell_obj.value='shift1'
		    cell_obj.fill = fill_cell1
		elif cell_obj.value=='ns':
		    cell_obj.value='shift2'
		    cell_obj.fill = fill_cell2
		print(cell_obj.value)
wrkbk.save("roster1.xlsx")
		    

